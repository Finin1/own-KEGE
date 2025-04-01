from pathlib import Path
from openpyxl import load_workbook, Workbook
from sqlalchemy import Select

from main import app
from word_parser import parse_Poliacov_document
from database import Student, create_session, create_db


def get_students_list(path_to_students_list: Path) -> bool:
    confirmation = input("Delete all old data? Y/N \n")
    if confirmation != "Y":
        exit(0)

    with create_session() as session:
        statement: Select = Select(Student)
        students = session.scalars(statement=statement).all()
        try:
            for student_to_remove in students:
                session.delete(student_to_remove)
            session.commit()
        except Exception as ex:
            print(ex)
            session.rollback()

    students_workbook = load_workbook(path_to_students_list)
    workbook_sheet = students_workbook.active

    row_iterator = workbook_sheet.iter_rows()

    new_students: list[Student] = []
    for name_cell, surname_cell in row_iterator:
        name: str = name_cell.internal_value
        surname: str = surname_cell.internal_value
        student: Student = Student(name=name, surname=surname)
        new_students.append(student)

    with create_session() as session:
        try:
            for student in new_students:
                session.add(student)
            session.commit()
        except Exception:
            session.rollback()
            return False

        workbook_with_codes = Workbook()
        active_sheet = workbook_with_codes.active
    
        for student in new_students:
            active_sheet.append([student.name, student.surname, student.code])
        
        workbook_with_codes.save("students_with_codes.xlsx")
    return True


if __name__ == "__main__":
    create_db()
    while not get_students_list(Path("students_list.xlsx")):
        pass
    parse_Poliacov_document(Path("test2.docx"))
    app.run(host="localhost", port=8080)
