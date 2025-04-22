import os
from pathlib import Path
from openpyxl import load_workbook, Workbook
try:
    from sqlalchemy import Select
except:
    from sqlalchemy import select as Select
from main import app
from word_parser import parse_Poliacov_document, parse_from_images
from database import Student, Task, StudentAnswer, create_session, create_db


POLIACOV_PARSE = '0'
IMAGES_PARSE = '1'


def parse_students_list(path_to_students_list: Path) -> bool:
    students_workbook = load_workbook(path_to_students_list)
    workbook_sheet = students_workbook.active

    row_iterator = workbook_sheet.iter_rows()

    new_students: list[Student] = []
    for name_cell, surname_cell in row_iterator:
        name: str = name_cell.internal_value
        surname: str = surname_cell.internal_value
        student: Student = Student(name=name, surname=surname)
        new_students.append(student)

    with create_session() as session:
        try:
            for student in new_students:
                session.add(student)
            session.commit()
        except Exception:
            session.rollback()
            return False

        workbook_with_codes = Workbook()
        active_sheet = workbook_with_codes.active

        for student in new_students:
            active_sheet.append([student.name, student.surname, student.code])

        workbook_with_codes.save("students_with_codes.xlsx")
    return True


# Нужно окно подтверждения
def flush_students() -> None:
    with create_session() as session:
        answers_statement = Select(StudentAnswer)
        answers = session.scalars(answers_statement).all()
        student_statement = Select(Student)
        students = session.scalars(student_statement).all()
        session.delete(*answers)
        session.delete(*students)
        session.commit()


def flush_tasks() -> None:
    with create_session() as session:
        tasks_statement = Select(Task)
        tasks = session.scalars(tasks_statement).all()
        session.delete(*tasks)
        session.commit()


def flush_database() -> None:
    flush_students()
    flush_tasks()        


def start_test() -> None:
    create_db()
    host = os.getenv("HOST", "localhost")
    port = int(os.getenv("PORT", 8080))
    app.run(host=host, port=port)


def init_test(path_to_root: Path, parse_type: str) -> None:
    flush_students()
    create_db()

    path_to_excel = path_to_root / "students_list.xlsx"
    path_to_word = path_to_root / "test2.docx"
    while not parse_students_list(path_to_excel):
        pass
    if parse_type == POLIACOV_PARSE:
        parse_Poliacov_document(path_to_word)
    elif parse_type == IMAGES_PARSE:
        parse_from_images(path_to_root)

    host = os.getenv("HOST", "localhost")
    port = int(os.getenv("PORT", 8080))
    app.run(host=host, port=port)


def get_score_results() -> None:
    conversion_scale = {0: 0, 1: 7, 2: 14, 3: 20, 4: 27,
                        5: 34, 6: 40, 7: 43, 8: 46, 9: 48,
                        10: 51, 11: 54, 12: 56, 13: 59, 14: 62,
                        15: 64, 16: 67, 17: 70, 18: 72, 19: 75,
                        20: 78, 21: 80, 22: 83, 23: 85, 24: 88,
                        25: 90, 26: 93, 27: 95, 28: 98, 29: 100}
    result_workbook = Workbook()
    active_sheet = result_workbook.active
    with create_session() as db_session:
        students_statement: Select = Select(Student)
        students = db_session.scalars(students_statement).all()
        task_statement: Select = Select(Task)
        tasks = db_session.scalars(task_statement).all()
        active_sheet.append(["Имя", "Фамилия"]
                            + [f"{task.task_number}\nТип: {task.task_type}" for task in tasks]
                            + ["Первичные баллы", "Вторичные баллы"])
        for student in students:
            name = student.name
            surname = student.surname
            new_row = [name, surname]
            new_row.extend([0] * len(tasks))
            number_to_row_number_dict = {}
            for i, task in enumerate(tasks, start=2):
                number_to_row_number_dict[task.task_number] = i
            total = 0
            answers = student.answers
            for answer in answers:
                task: Task = answer.task
                number: int = task.task_number
                task_type: int = task.task_type
                correct_answer: str = task.task_answer
                student_answer: str = answer.student_answer
                if student_answer == "None":
                    new_row[number_to_row_number_dict[number]] = "-"
                if task_type == 26:
                    splitted_student_answer = student_answer.split()
                    splitted_correct_answer = correct_answer.split()
                    new_row[number_to_row_number_dict[number]] = 0
                    if splitted_student_answer[0] == splitted_correct_answer[0]:
                        new_row[number_to_row_number_dict[number]] += 1
                        total += 1
                    if splitted_student_answer[1] == splitted_correct_answer[1]:
                        new_row[number_to_row_number_dict[number]] += 1
                        total += 1
                elif task_type == 27:
                    splitted_student_answer = student_answer.split("\n")
                    splitted_correct_answer = correct_answer.split("\n")
                    new_row[number_to_row_number_dict[number]] = 0
                    if splitted_student_answer[0] == splitted_correct_answer[0]:
                        new_row[number_to_row_number_dict[number]] += 1
                        total += 1
                    if splitted_student_answer[1] == splitted_correct_answer[1]:
                        new_row[number_to_row_number_dict[number]] += 1
                        total += 1
                else:
                    if student_answer == correct_answer:
                        new_row[number_to_row_number_dict[number]] = 1
                        total += 1
                    else:
                        new_row[number_to_row_number_dict[number]] = 0
            new_row.append(total)
            new_row.append(conversion_scale[total])
            active_sheet.append(new_row)
        result_workbook.save("results.xlsx")




if __name__ == "__main__":
    inp = input("Init or result?\n")
    if inp == "init":
        init_test(Path(""))
    elif inp == "result":
        get_score_results()
